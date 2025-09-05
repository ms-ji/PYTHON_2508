from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


def main():
    """
    
    """
    # 엑셀 파일 열기
    wb = load_workbook('address_book.xlsx')

    #시트 선택
    ws = wb.active

    #데이터 읽기
    for row in ws.iter_rows(values_only=True):
        print(f'row: {row}')

    #셀 값 읽기
    print(f'ws[\'A2\'].value:{ws['A2'].value}') # 영희
    print(f'ws.cell(2,2).value:{ws.cell(2,2).value}') #22
    
    #셀 값 쓰기
    ws['C1'] = '점수'
    ws['C2'] = 70
    ws['C3'] = 75
 
    
    #저장
    wb.save('new_address.xlsx')
    print('저장 완료')


if __name__ == '__main__':
    main()
