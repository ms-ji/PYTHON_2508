from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
import re

def fine():
    print(f'fine()')
    #크롤링
    list_titles = [] # 용어
    list_contents = [] # 용어 설명

    fDic = {} # 용어 설명

    try:
        #for i in range(1, 54 + 1): 운영 시에 바꾸면 됨.
        for i in range(1,54+1):
            pass
            #print(f'i:{i}',end=",")
            url = f'https://fine.fss.or.kr/fine/fnctip/fncDicary/list.do?menuNo=900021&src=&kind=&searchCnd=1&searchStr=&pageIndex={i}'
            #print(f'url: {url}')

            #웹 사이트 속성
            response = requests.get(url)

            if response.status_code == 200:
                #print(f'text\n: {response.text}')
                html =  response.text
                #bs4로 parsing
                bs = BeautifulSoup(html,'html.parser')

                #용어 class="bd-list result-list"
                titles = bs.select('div.result-list dl dt')

                for title in titles:
                    print(f'title:{title.text}')
                    cleaned_title = re.sub(r'[^가-힣0-9.a-zA-Z\[\]]', '', title.text)
                    list_titles.append(cleaned_title)

                print(f'list_titles: {list_titles}')

                #용어 설명
                contents = bs.select('div.result-list dl dd')
                for content in contents:
                    print(f'content:{content.text}')
                    cleaned_content = re.sub(r'[\\n]', '', content.text)
                    list_contents.append(cleaned_content)
                print(f'list_contents: {list_contents}')

                print(f'titles_len:{len(list_titles)}')
                print(f'contents_len:{len(list_contents)}')

                # 용어 + 용어 사전 => Dic(용어: 용어 사전)
                for i in range(0,len(list_titles)):
                    print(f'i: {i}',end=",")
                    fDic[list_titles[i]] = list_contents[i]
                print(f'Dic:\n{fDic}')



            else:
                print(f'접속 실패:{response.status_code}')

    except ConnectionError as e:
        print(f'ConnectionError(접속 오류)+{e}')
    except Exception as e:
        print(f'Exception(알 수 없는 오류)+{e}')
    else:
        print("*"*80)
        print(f'정상적으로 프로그램이 종료 되었습니다.')
        print("*" * 80)
    finally:
        print('항상 수행')
    return fDic

def writeXlsx(fineDic):
    print(f'writeXlsx fineDic:{fineDic}')

    #새 워크북 생성
    wb=Workbook()

    #시트 선택
    ws = wb.active

    #시트 이름 지정
    ws.title ='fine_금융용어사전'

    #데이터 쓰기
    ws['A1'] = '금융 용어'
    ws['B1'] = '용어 사전'

    for key in fineDic.keys():
        ws.append([key, fineDic.get(key)])
    # 파일에 저장
    wb.save('금융용어사전_20250807.xlsx')
    print('금융용어사전_20250807.xlsx 파일 저장 성공')

def main():
    """
    
    """
    #fine 크롤링
    fineDic = fine()

    #엑셀 저장
    writeXlsx(fineDic)



if __name__ == '__main__':
    main()
