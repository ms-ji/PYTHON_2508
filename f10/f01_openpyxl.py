from openpyxl import Workbook

def main():
    """
    
    """
    # 새 워크북 생성
    wb = Workbook()

    #시트
    ws = wb.active
    
    ws.title = 'PCWK_403_Sheet'
    
    #데이터 쓰기
    ws['A1'] = '이름'
    ws['B1'] = '나이'
    ws.append(['영희',22]) #행 추가

    #파일 저장
    wb.save('address_book.xlsx')
    print('저장되었습니다.')


if __name__ == '__main__':
    main()
